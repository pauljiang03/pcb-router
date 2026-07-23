"""Board representation for SI test fixture routing: geometry, obstacles,
start points, candidate grid generation, and constraint checks."""

import numpy as np
from dataclasses import dataclass, field
from typing import List, Tuple, Optional


@dataclass
class Obstacle:
    """Rectangular obstacle or keep-out zone."""
    cx: float
    cy: float
    width: float
    height: float
    clearance: float  # edge-to-edge
    name: str = ""

    @property
    def bounds(self) -> Tuple[float, float, float, float]:
        """Return (x_min, y_min, x_max, y_max)."""
        hw, hh = self.width / 2, self.height / 2
        return (self.cx - hw, self.cy - hh, self.cx + hw, self.cy + hh)


@dataclass
class CircularObstacle:
    """Circular obstacle (UPTH, via, etc.)."""
    cx: float
    cy: float
    radius: float
    clearance: float  # edge-to-edge
    name: str = ""


@dataclass
class TraceSpec:
    """Specification for one trace to be routed."""
    start_x: float
    start_y: float
    breakout_length: float
    index: int


@dataclass
class BoardSpec:
    """Complete board specification loaded from data."""
    origin_x: float
    origin_y: float
    width: float
    height: float

    rect_obstacles: List[Obstacle] = field(default_factory=list)
    circ_obstacles: List[CircularObstacle] = field(default_factory=list)
    traces: List[TraceSpec] = field(default_factory=list)

    # Connector outline (rectangle)
    connector_x: float = 0.0
    connector_y: float = 0.0
    connector_w: float = 0.0
    connector_h: float = 0.0

    @property
    def x_min(self):
        return self.origin_x

    @property
    def y_min(self):
        return self.origin_y

    @property
    def x_max(self):
        return self.origin_x + self.width

    @property
    def y_max(self):
        return self.origin_y + self.height


# Fixed constraints (TE AutoLayout Example01)
TRACE_WIDTH = 0.2286        # mm
TRACE_TO_EDGE_MIN = 0.26    # mm, edge-to-edge
TRACE_TO_TRACE_MIN = 1.1    # mm, edge-to-edge
TRACE_TO_UPTH_MIN = 0.7     # mm, edge-to-edge
TRACE_TO_TABPAD_MIN = 0.7   # mm, edge-to-edge
TP_TO_TP_MIN = 13.0         # mm, center-to-center
TP_TO_EDGE_MIN = 14.0       # mm, center-to-edge
TP_TO_CONNECTOR_MIN = 3.0   # mm, center-to-edge

TRACE_MIN_CENTER_TO_CENTER = TRACE_TO_TRACE_MIN + TRACE_WIDTH  # 1.3286mm

# Fixed action-space size, constant across board geometries.
MAX_CANDIDATES = 200

BOARD_SIZE = 160.0          # square board edge, mm
N_PER_ROW = 10              # starts per row; 2 rows -> 20 traces


def _respaced_x(original_x: List[float], min_spacing: float) -> List[float]:
    """Widen only sub-minimum gaps in an ascending x list to min_spacing, preserving the mean."""
    n = len(original_x)
    if n <= 1:
        return list(original_x)
    new_x = [0.0]
    for a, b in zip(original_x, original_x[1:]):
        new_x.append(new_x[-1] + max(b - a, min_spacing))
    off = float(np.mean(original_x)) - float(np.mean(new_x))
    return [v + off for v in new_x]


def load_te_example(num_traces: int = 20, seed: int = None,
                    board_size: float = BOARD_SIZE) -> BoardSpec:
    """Central-connector board: two pin rows straddle the NRZ at board center (lower row first)."""
    board = BoardSpec(origin_x=0.0, origin_y=0.0,
                      width=board_size, height=board_size)

    cx0 = board.x_min + board.width / 2.0
    cy0 = board.y_min + board.height / 2.0

    min_sp = TRACE_MIN_CENTER_TO_CENTER
    row_span = (N_PER_ROW - 1) * min_sp

    nrz_w = row_span + 4.0
    nrz_h = 6.64
    row_gap = nrz_h + 0.2

    if seed is not None:
        rng = np.random.RandomState(seed)
        jitter = min(20.0, board.width / 2 - (TP_TO_EDGE_MIN + max(nrz_w, row_gap)))
        ccx = cx0 + rng.uniform(-jitter, jitter)
        ccy = cy0 + rng.uniform(-jitter, jitter)
    else:
        ccx, ccy = cx0, cy0

    board.rect_obstacles.append(Obstacle(
        cx=ccx, cy=ccy, width=nrz_w, height=nrz_h,
        clearance=TRACE_TO_EDGE_MIN, name="non_routing_zone",
    ))

    board.circ_obstacles.append(CircularObstacle(
        cx=ccx - nrz_w * 0.25, cy=ccy, radius=1.9 / 2,
        clearance=TRACE_TO_UPTH_MIN, name="UPTH_1"))
    board.circ_obstacles.append(CircularObstacle(
        cx=ccx + nrz_w * 0.25, cy=ccy, radius=1.9 / 2,
        clearance=TRACE_TO_UPTH_MIN, name="UPTH_2"))

    for sgn, nm in ((-1, "tab_pad_1"), (1, "tab_pad_2")):
        board.rect_obstacles.append(Obstacle(
            cx=ccx + sgn * (row_span / 2 + 1.2), cy=ccy,
            width=1.526, height=1.216,
            clearance=TRACE_TO_TABPAD_MIN, name=nm))

    conn_w = row_span + 6.0
    conn_h = row_gap + 6.0
    board.connector_x = ccx - conn_w / 2
    board.connector_y = ccy - conn_h / 2
    board.connector_w = conn_w
    board.connector_h = conn_h

    start_xs = [ccx + (i - (N_PER_ROW - 1) / 2.0) * min_sp
                for i in range(N_PER_ROW)]
    lower_y = ccy - nrz_h / 2 - 0.1               # escapes downward
    upper_y = ccy + nrz_h / 2 + 0.1               # escapes upward
    breakout = 0.8626

    all_traces = []
    for i, x in enumerate(start_xs):              # lower row: 0 .. N-1
        all_traces.append(TraceSpec(start_x=x, start_y=lower_y,
                                    breakout_length=breakout, index=i))
    for i, x in enumerate(start_xs):              # upper row: N .. 2N-1
        all_traces.append(TraceSpec(start_x=x, start_y=upper_y,
                                    breakout_length=breakout, index=N_PER_ROW + i))

    board.traces = all_traces[:num_traces]
    return board


def load_te_excel(path: Optional[str] = None, respace: bool = True,
                  center_on_nrz: bool = True) -> BoardSpec:
    """Load the canonical TE AutoLayout Example01 board from its Excel file (needs openpyxl)."""
    import pathlib
    import openpyxl
    if path is None:
        path = str(pathlib.Path(__file__).resolve().parents[1]
                   / "AutoLayout_Example01.xlsx")
    # data_only=True: read cached formula VALUES, not formula strings.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    def block_at(i):
        """Parse block at row i: ('x'|'y', value) in cols E/F, dimensions in cols D/E."""
        vals = {}
        j = i
        while j < len(rows):
            r = rows[j]
            if j > i and r[2] is not None:        # next section begins
                break
            if r[4] in ("x", "y"):
                vals[r[4]] = float(r[5])
            elif isinstance(r[3], str) and r[4] is not None:
                vals[r[3]] = float(r[4])
            j += 1
        return vals

    board_v, nrz, upths, tabs, pins = None, None, [], [], []
    for i, r in enumerate(rows):
        tag = r[2]
        if tag is None:
            continue
        if tag == "Size":
            board_v = block_at(i)
        elif tag == "Non-routing zone":
            nrz = block_at(i)
        elif isinstance(tag, str) and tag.startswith("UPTH"):
            upths.append(block_at(i))
        elif isinstance(tag, str) and tag.startswith("Tab Pad"):
            tabs.append(block_at(i))
        elif isinstance(tag, (int, float)) and r[3] == "Starting Point":
            v = block_at(i)
            pins.append((int(tag), v["x"], v["y"], v.get("Routed trace length", 0.0)))

    assert board_v and nrz and pins, f"unrecognized board file: {path}"
    board = BoardSpec(origin_x=board_v["x"], origin_y=board_v["y"],
                      width=board_v["Width"], height=board_v["Height"])
    board.rect_obstacles.append(Obstacle(
        cx=nrz["x"] + nrz["Width"] / 2, cy=nrz["y"] + nrz["Height"] / 2,
        width=nrz["Width"], height=nrz["Height"],
        clearance=TRACE_TO_EDGE_MIN, name="non_routing_zone"))
    for k, u in enumerate(upths):
        board.circ_obstacles.append(CircularObstacle(
            cx=u["x"], cy=u["y"], radius=u["Diameter"] / 2,
            clearance=TRACE_TO_UPTH_MIN, name=f"UPTH_{k + 1}"))
    for k, t in enumerate(tabs):
        board.rect_obstacles.append(Obstacle(
            cx=t["x"] + t["Width"] / 2, cy=t["y"] + t["Height"] / 2,
            width=t["Width"], height=t["Height"],
            clearance=TRACE_TO_TABPAD_MIN, name=f"tab_pad_{k + 1}"))

    # Widen each pin row to the minimum legal pitch if needed, then center it on the NRZ.
    pins.sort(key=lambda p: p[0])                  # file numbering 1..20
    nrz_cx = nrz["x"] + nrz["Width"] / 2
    xs_of = {}
    by_row = {}
    for num, x, y, bl in pins:
        by_row.setdefault(round(y, 3), []).append((num, x))
    for yk, items in by_row.items():
        items.sort(key=lambda t: t[1])
        xs = [x for _n, x in items]
        if respace and len(xs) > 1 and \
                np.diff(xs).min() < TRACE_MIN_CENTER_TO_CENTER - 1e-9:
            xs = _respaced_x(xs, TRACE_MIN_CENTER_TO_CENTER)
        if center_on_nrz:
            shift = nrz_cx - (xs[0] + xs[-1]) / 2
            xs = [x + shift for x in xs]
        for (num, _x), nx in zip(items, xs):
            xs_of[num] = nx
    for num, x, y, bl in pins:
        board.traces.append(TraceSpec(start_x=xs_of.get(num, x), start_y=y,
                                      breakout_length=bl, index=num - 1))

    # Connector outline = cluster bounding box + margin (used for TP_TO_CONNECTOR_MIN).
    ex = [o.bounds[0] for o in board.rect_obstacles] + \
         [o.bounds[2] for o in board.rect_obstacles] + \
         [u.cx - u.radius for u in board.circ_obstacles] + \
         [u.cx + u.radius for u in board.circ_obstacles] + \
         [t.start_x for t in board.traces]
    ey = [o.bounds[1] for o in board.rect_obstacles] + \
         [o.bounds[3] for o in board.rect_obstacles] + \
         [u.cy - u.radius for u in board.circ_obstacles] + \
         [u.cy + u.radius for u in board.circ_obstacles] + \
         [t.start_y for t in board.traces]
    m = 1.5
    board.connector_x, board.connector_y = min(ex) - m, min(ey) - m
    board.connector_w = max(ex) - min(ex) + 2 * m
    board.connector_h = max(ey) - min(ey) + 2 * m
    return board


def _octd(ax, ay, bx, by):
    dx, dy = abs(ax - bx), abs(ay - by)
    return max(dx, dy) + 0.4142135624 * min(dx, dy)


def wire_estimate(board: BoardSpec, pin, pad) -> float:
    """Realizable wire-length estimate pin->pad: octile distance, routed around the NRZ if crossed."""
    px, py = (pin.start_x, pin.start_y) if hasattr(pin, "start_x") else pin
    qx, qy = pad
    nrz = next((o for o in board.rect_obstacles
                if o.name == "non_routing_zone"), None)
    if nrz is None:
        return _octd(px, py, qx, qy)
    # Crossing test uses the actual rect shrunk by eps (pins sit ~0.1mm outside; any
    # inflation would swallow them); corner nodes sit clearance + one lane outside.
    bx0, by0, bx1, by1 = nrz.bounds
    eps = 1e-6
    x0, y0, x1, y1 = bx0 + eps, by0 + eps, bx1 - eps, by1 - eps
    m = TRACE_TO_EDGE_MIN + TRACE_WIDTH + TRACE_TO_TRACE_MIN
    cx0, cy0, cx1, cy1 = bx0 - m, by0 - m, bx1 + m, by1 + m

    def crosses(ax, ay, bx_, by_):
        """Does segment a->b pass through the open NRZ interior?"""
        dx, dy = bx_ - ax, by_ - ay
        t0, t1 = 0.0, 1.0
        for p_, q_ in ((-dx, ax - x0), (dx, x1 - ax), (-dy, ay - y0), (dy, y1 - ay)):
            if abs(p_) < 1e-12:
                if q_ < 0:
                    return False
                continue
            r = q_ / p_
            if p_ < 0:
                t0 = max(t0, r)
            else:
                t1 = min(t1, r)
            if t0 - t1 > 1e-12:
                return False
        return t1 - t0 > 1e-9

    nodes = [(px, py), (qx, qy), (cx0, cy0), (cx1, cy0), (cx0, cy1), (cx1, cy1)]
    if not crosses(*nodes[0], *nodes[1]):
        return _octd(px, py, qx, qy)
    INF = float("inf")
    dist = [INF] * 6
    dist[0] = 0.0
    done = [False] * 6
    for _ in range(6):
        u = min((d, i) for i, d in enumerate(dist) if not done[i])[1]
        done[u] = True
        if u == 1:
            break
        for v in range(6):
            if not done[v] and not crosses(*nodes[u], *nodes[v]):
                nd = dist[u] + _octd(*nodes[u], *nodes[v])
                if nd < dist[v]:
                    dist[v] = nd
    return dist[1] if dist[1] < INF else _octd(px, py, qx, qy) * 3.0


def repair_placement(board: BoardSpec, placed, rounds: int = 2,
                     k_alt: int = 12, verbose: bool = False,
                     objective: str = "fails"):
    """Relocate problem-net pads to nearby legal spots, keeping moves that improve the objective."""
    from envs.routing import route_all_traces, equalize_lengths
    placed = list(placed)
    cand, rc = generate_candidate_grid(board, 6.5)
    cand = cand[:rc]
    fast = dict(n_starts=1, max_iters=12, repair_passes=1)

    def measure():
        paths, _L, fails = route_all_traces(board, placed, **fast)
        if objective == "fails":
            return paths, (fails,), []
        _eq, eqL, tgt, _m = equalize_lengths(board, paths, test_points=placed)
        fin = [x for x in eqL if x < 1e9]
        target = max(fin) if fin else 0.0
        bad = [i for i in range(len(placed))
               if paths[i] is not None and eqL[i] < target - 1.5]
        problem = bad if fails == 0 else \
            [i for i in range(len(placed)) if paths[i] is None]
        return paths, (fails, len(bad)), problem

    paths, score, problem = measure()
    for _ in range(rounds):
        if (objective == "fails" and score[0] == 0) or \
           (objective == "matched" and score == (0, 0)):
            break
        if objective == "fails":
            problem = [j for j in range(len(placed)) if paths[j] is None]
        improved = False
        for i in problem:
            others = [placed[j] for j in range(len(placed))
                      if j != i and placed[j] is not None]
            d = np.hypot(cand[:, 0] - placed[i][0], cand[:, 1] - placed[i][1])
            tried = 0
            for idx in np.argsort(d):
                if tried >= k_alt:
                    break
                p = tuple(cand[idx])
                if p == placed[i]:
                    continue
                if not all(np.hypot(p[0] - q[0], p[1] - q[1]) >= TP_TO_TP_MIN
                           for q in others):
                    continue
                tried += 1
                old = placed[i]
                placed[i] = p
                p2, s2, prob2 = measure()
                if s2 < score:
                    score, paths, problem = s2, p2, prob2
                    improved = True
                    if verbose:
                        print(f"  repair[{objective}]: net {i} pad -> "
                              f"({p[0]:.1f},{p[1]:.1f}) score {s2}")
                    break
                placed[i] = old
            if (objective == "fails" and score[0] == 0) or \
               (objective == "matched" and score == (0, 0)):
                break
        if not improved:
            break
    return placed


def smart_placement(board: BoardSpec, num_traces: int, elect: bool = True,
                    verbose: bool = False):
    """No-training placement baseline: elect the best candidate pad set by trial routing."""
    cands = _smart_candidates(board, num_traces)
    if not elect:
        return cands["near"]
    from envs.routing import route_all_traces, equalize_lengths
    best_name, best_key, best = None, None, None
    for name, placed in cands.items():
        paths, lengths, fails = route_all_traces(
            board, placed, n_starts=1, max_iters=12, repair_passes=1)
        if 0 < fails <= 3:
            # Repair near-miss candidates before judging.
            placed = repair_placement(board, placed, verbose=verbose)
            paths, lengths, fails = route_all_traces(
                board, placed, n_starts=1, max_iters=12, repair_passes=1)
        _eq, _eqL, _t, matched = equalize_lengths(
            board, paths, test_points=placed)
        fin = [l for l in lengths if l < float("inf")]
        unmatched = (len(fin) - matched)
        key = (fails, unmatched, max(fin) if fin else 0.0, sum(fin))
        if verbose:
            print(f"  smart_placement[{name}]: fails={fails} "
                  f"unmatched={unmatched} max={key[2]:.0f}mm "
                  f"total={key[3]:.0f}mm")
        if best_key is None or key < best_key:
            best_name, best_key, best = name, key, placed
    if verbose:
        print(f"  smart_placement elected: {best_name}")
    if best_key[0] == 0 and best_key[1] > 0:
        # Fix unmatched nets without trading away a routing success.
        best = repair_placement(board, best, objective="matched",
                                verbose=verbose)
    return best


def _smart_candidates(board: BoardSpec, num_traces: int):
    """The candidate placements smart_placement elects among (name -> list)."""
    cand, real = generate_candidate_grid(board, 6.5)
    cand = cand[:real]
    ccx = board.connector_x + board.connector_w / 2
    ccy = board.connector_y + board.connector_h / 2
    dist = np.hypot(cand[:, 0] - ccx, cand[:, 1] - ccy)

    def pick(order, min_sp=None):
        """Greedy spacing-valid pick; min-spaced pads pinch corridors, so callers may ask larger min_sp."""
        for sp in ([min_sp] if min_sp else [None]):
            chosen = []
            for idx in order:
                if len(chosen) >= num_traces:
                    break
                p = tuple(cand[idx])
                if sp is None:
                    ok = check_tp_spacing(chosen, *p)
                else:
                    ok = all(np.hypot(p[0] - q[0], p[1] - q[1]) >= sp
                             for q in chosen)
                if ok:
                    chosen.append(p)
            if len(chosen) >= num_traces:
                return chosen
        return chosen

    def pick_airy(order):
        """Largest spacing that still fits num_traces pads."""
        from envs.routing import TP_CLEARANCE_CELLS, CELL_SIZE
        lane = CELL_SIZE
        airy = 2 * TP_CLEARANCE_CELLS * CELL_SIZE + 3 * lane  # disks + lanes
        for sp in (max(airy, TP_TO_TP_MIN), TP_TO_TP_MIN + 2.0, TP_TO_TP_MIN):
            chosen = pick(order, min_sp=sp)
            if len(chosen) >= num_traces:
                return chosen
        return pick(order)

    # Pin->pad cost = realizable wire length; air distance ignores the NRZ wrap and mis-assigns.
    _wcache: dict = {}

    def wcost(i, p):
        key = (i, p)
        if key not in _wcache:
            _wcache[key] = wire_estimate(board, board.traces[i], p)
        return _wcache[key]

    def match(chosen):
        """Angle-sorted pins x pads at the cyclic offset with minimum total realizable wire."""
        m = len(chosen)
        if m == 0:
            return [None] * num_traces
        tps = sorted(chosen, key=lambda p: np.arctan2(p[1] - ccy, p[0] - ccx))
        pins = sorted(range(num_traces),
                      key=lambda i: np.arctan2(board.traces[i].start_y - ccy,
                                               board.traces[i].start_x - ccx))
        best_off, best_cost = 0, None
        for off in range(m):
            cost = sum(wcost(i, tps[(k + off) % m])
                       for k, i in enumerate(pins))
            if best_cost is None or cost < best_cost:
                best_off, best_cost = off, cost
        placed = [None] * num_traces
        for k, i in enumerate(pins):
            placed[i] = tps[(k + best_off) % m]
        return placed

    def eqwire():
        """Equal-wire ring: pads at a common realizable wire length, smallest matched spread."""
        emin = np.array([min(wire_estimate(board, t, tuple(c))
                             for t in board.traces[:num_traces])
                         for c in cand])
        best_pl, best_spread = None, None
        for q in (0.35, 0.45, 0.55, 0.65, 0.75):
            R = float(np.quantile(emin, q))
            pl = match(pick_airy(np.argsort(np.abs(emin - R))))
            ests = [wcost(i, pl[i]) for i in range(num_traces) if pl[i]]
            if len(ests) < num_traces:
                continue
            spr = (max(ests) - min(ests)) / max(np.mean(ests), 1e-9)
            if best_spread is None or spr < best_spread:
                best_pl, best_spread = pl, spr
        return best_pl if best_pl is not None else match(pick_airy(np.argsort(dist)))

    ring = equal_length_placement(board, num_traces)
    spread = spread_placement(board, num_traces)
    return {
        "near": match(pick_airy(np.argsort(dist))),
        # Re-match ring/spread pad sets: their built-in matching can use a pathological cut.
        "ring": match([p for p in ring if p is not None]),
        "spread": match([p for p in spread if p is not None]),
        "eqwire": eqwire(),
        "airy": match(pick_airy(np.argsort(-dist))),
    }


def load_edge_board(num_traces: int = 20, board_w: float = 180.0,
                    board_h: float = 180.0, seed: int = None) -> BoardSpec:
    """Edge-connector board: one pin row low on the board, all escaping upward."""
    board = BoardSpec(origin_x=0.0, origin_y=0.0, width=board_w, height=board_h)
    cx = board_w / 2.0
    if seed is not None:
        cx += np.random.RandomState(seed).uniform(-board_w * 0.1, board_w * 0.1)
    row_y = board_h * 0.12
    sp = TRACE_MIN_CENTER_TO_CENTER
    xs = [cx + (i - (num_traces - 1) / 2) * sp for i in range(num_traces)]
    span = (num_traces - 1) * sp
    # NRZ + UPTH sit below the pin row, so traces must escape upward.
    board.rect_obstacles.append(Obstacle(
        cx=cx, cy=row_y - 6.0, width=span + 4.0, height=7.0,
        clearance=TRACE_TO_EDGE_MIN, name="non_routing_zone"))
    board.circ_obstacles.append(CircularObstacle(
        cx=cx, cy=row_y - 6.0, radius=0.95, clearance=TRACE_TO_UPTH_MIN, name="UPTH_1"))
    board.connector_x = cx - span / 2 - 3.0
    board.connector_y = row_y - 10.0
    board.connector_w = span + 6.0
    board.connector_h = 13.0
    board.traces = [TraceSpec(start_x=x, start_y=row_y, breakout_length=0.8626, index=i)
                    for i, x in enumerate(xs)]
    return board


def fan_to_top_placement(board: BoardSpec, num_traces: int,
                         rows: int = 2) -> List[Tuple[float, float]]:
    """Planar fan to the top for load_edge_board: TP rows up high, matched to pins by x."""
    per = -(-num_traces // rows)
    xlo = board.x_min + TP_TO_EDGE_MIN + 5.0
    xhi = board.x_max - TP_TO_EDGE_MIN - 5.0
    top = board.y_max - TP_TO_EDGE_MIN
    start_top = max(t.start_y for t in board.traces)
    rowys = np.linspace(start_top + 0.55 * (top - start_top), top, rows)
    xs = np.linspace(xlo, xhi, per)
    tps = sorted([(float(x), float(y)) for y in rowys for x in xs],
                 key=lambda p: p[0])[:num_traces]
    pins = sorted(range(num_traces), key=lambda i: board.traces[i].start_x)
    placed = [None] * num_traces
    for k, i in enumerate(pins):
        placed[i] = tps[k] if k < len(tps) else tps[-1]
    return placed


def load_edge_board_2row(num_traces: int = 20, board_w: float = 240.0,
                         board_h: float = 250.0, conn_yf: float = 0.38,
                         seed: int = None) -> BoardSpec:
    """Two-row edge connector low on the board: lower row wraps up the sides, upper escapes up."""
    board = BoardSpec(origin_x=0.0, origin_y=0.0, width=board_w, height=board_h)
    cx = board_w / 2.0
    if seed is not None:
        cx += np.random.RandomState(seed).uniform(-board_w * 0.06, board_w * 0.06)
    ccy = board_h * conn_yf
    per = num_traces // 2
    sp = TRACE_MIN_CENTER_TO_CENTER
    xs = [cx + (i - (per - 1) / 2) * sp for i in range(per)]
    span = (per - 1) * sp
    nrz_h = 6.64
    board.rect_obstacles.append(Obstacle(
        cx=cx, cy=ccy, width=span + 4.0, height=nrz_h,
        clearance=TRACE_TO_EDGE_MIN, name="non_routing_zone"))
    board.circ_obstacles.append(CircularObstacle(
        cx=cx, cy=ccy, radius=0.95, clearance=TRACE_TO_UPTH_MIN, name="UPTH_1"))
    board.connector_x = cx - span / 2 - 3.0
    board.connector_y = ccy - (nrz_h / 2 + 3.0)
    board.connector_w = span + 6.0
    board.connector_h = nrz_h + 6.0
    lower_y = ccy - nrz_h / 2 - 0.1          # escapes downward
    upper_y = ccy + nrz_h / 2 + 0.1          # escapes upward
    traces = [TraceSpec(start_x=x, start_y=lower_y, breakout_length=0.8626, index=i)
              for i, x in enumerate(xs)]
    traces += [TraceSpec(start_x=x, start_y=upper_y, breakout_length=0.8626, index=per + i)
               for i, x in enumerate(xs)]
    board.traces = traces[:num_traces]
    return board


def wrap_to_top_placement(board: BoardSpec, num_traces: int) -> List[Tuple[float, float]]:
    """Planar fan to the top for load_edge_board_2row: lower row to outer TPs, upper to middle."""
    per = num_traces // 2
    ccy = board.connector_y + board.connector_h / 2
    xlo = board.x_min + TP_TO_EDGE_MIN + 5.0
    xhi = board.x_max - TP_TO_EDGE_MIN - 5.0
    top = board.y_max - TP_TO_EDGE_MIN
    rowys = np.linspace(ccy + 0.55 * (top - ccy), top, 2)
    xs = np.linspace(xlo, xhi, (num_traces + 1) // 2)
    tps = sorted([(float(x), float(y)) for y in rowys for x in xs],
                 key=lambda p: p[0])[:num_traces]
    lower = sorted(range(per), key=lambda i: board.traces[i].start_x)
    upper = sorted(range(per, num_traces), key=lambda i: board.traces[i].start_x)
    h = per // 2
    placed = [None] * num_traces
    for k in range(h):
        placed[lower[k]] = tps[k]                         # lower-left -> leftmost
    for k in range(per):
        placed[upper[k]] = tps[h + k]                     # upper -> middle
    for k in range(per - h):
        placed[lower[h + k]] = tps[h + per + k]           # lower-right -> rightmost
    return placed


def spread_placement(board: BoardSpec, num_traces: int):
    """Non-crossing radial fan: well-separated TPs around the connector, matched to pins by angle."""
    cand, real = generate_candidate_grid(board, 6.5)
    cand = cand[:real]
    ccx = board.connector_x + board.connector_w / 2
    ccy = board.connector_y + board.connector_h / 2
    chosen = []
    for idx in np.argsort(-np.hypot(cand[:, 0] - ccx, cand[:, 1] - ccy)):
        if len(chosen) >= num_traces:
            break
        if check_tp_spacing(chosen, *cand[idx]):
            chosen.append(tuple(cand[idx]))
    tps = sorted(chosen[:num_traces], key=lambda p: np.arctan2(p[1] - ccy, p[0] - ccx))
    pins = sorted(range(num_traces),
                  key=lambda i: np.arctan2(board.traces[i].start_y - ccy,
                                           board.traces[i].start_x - ccx))
    placed = [None] * num_traces
    for k, i in enumerate(pins):
        placed[i] = tps[k] if k < len(tps) else tps[-1]
    return placed


def equal_length_placement(board: BoardSpec, num_traces: int):
    """TPs on a common-radius ring around the connector, matched by angle for near-equal lengths."""
    cand, real = generate_candidate_grid(board, 6.5)
    cand = cand[:real]
    ccx = board.connector_x + board.connector_w / 2
    ccy = board.connector_y + board.connector_h / 2
    dist = np.hypot(cand[:, 0] - ccx, cand[:, 1] - ccy)
    R = float(np.median(dist))
    chosen = []
    for idx in np.argsort(np.abs(dist - R)):
        if len(chosen) >= num_traces:
            break
        if check_tp_spacing(chosen, *cand[idx]):
            chosen.append(tuple(cand[idx]))
    tps = sorted(chosen, key=lambda p: np.arctan2(p[1] - ccy, p[0] - ccx))
    pins = sorted(range(num_traces),
                  key=lambda i: np.arctan2(board.traces[i].start_y - ccy,
                                           board.traces[i].start_x - ccx))
    placed = [None] * num_traces
    for k, i in enumerate(pins):
        placed[i] = tps[k] if k < len(tps) else tps[-1]
    return placed


@dataclass
class ChallengeSpec:
    """Parametric 'moat' challenge board: obstacle ring around the connector with n_gaps gaps."""
    board_size: float = 120.0
    num_traces: int = 20
    n_gaps: int = 3
    gap_halfwidth: float = 0.45        # radians
    moat_radius_frac: float = 0.30     # ring radius / board_size
    obstacle_frac: float = 0.055       # obstacle side / board_size
    moat_segments: int = 22            # obstacle slots around the ring
    tp_radius_mult: float = 1.25       # pads beyond moat_radius * this
    seed: int = 6
    placement: str = "ring"            # "ring" | "gap_aligned"
    assignment: str = "angle"          # "angle" | "gap_aware"


def make_challenge(spec: ChallengeSpec):
    """Build a parametric moat board + a pad placement; returns (board, placed)."""
    n, ng = spec.num_traces, spec.n_gaps
    board = load_te_example(num_traces=n, seed=spec.seed, board_size=spec.board_size)
    cx = board.connector_x + board.connector_w / 2
    cy = board.connector_y + board.connector_h / 2
    R = spec.board_size * spec.moat_radius_frac
    gaps = [2 * np.pi * g / ng for g in range(ng)]
    side = spec.board_size * spec.obstacle_frac
    for i in range(spec.moat_segments):
        a = 2 * np.pi * i / spec.moat_segments
        if any(abs(((a - g + np.pi) % (2 * np.pi)) - np.pi) < spec.gap_halfwidth for g in gaps):
            continue
        board.rect_obstacles.append(Obstacle(
            cx=cx + R * np.cos(a), cy=cy + R * np.sin(a),
            width=side, height=side, clearance=TRACE_TO_TRACE_MIN, name="moat"))
    cand, real = generate_candidate_grid(board, 6.5)
    cand = cand[:real]
    out = cand[np.hypot(cand[:, 0] - cx, cand[:, 1] - cy) > R * spec.tp_radius_mult]
    ang = lambda x, y: np.arctan2(y - cy, x - cx)
    near_gap = lambda x, y: min(range(ng),
        key=lambda g: abs(((ang(x, y) - gaps[g] + np.pi) % (2 * np.pi)) - np.pi))

    if spec.placement == "gap_aligned":
        oa = np.arctan2(out[:, 1] - cy, out[:, 0] - cx)
        per = [n // ng + (1 if g < n % ng else 0) for g in range(ng)]   # pads per gap
        chosen = []
        for g in range(ng):
            angd = np.abs(((oa - gaps[g] + np.pi) % (2 * np.pi)) - np.pi)
            cnt = 0
            for idx in np.argsort(angd):
                if cnt >= per[g]:
                    break
                if check_tp_spacing(chosen, *out[idx]):
                    chosen.append(tuple(out[idx])); cnt += 1
        for idx in np.argsort(-np.hypot(out[:, 0] - cx, out[:, 1] - cy)):  # fill remainder
            if len(chosen) >= n:
                break
            if check_tp_spacing(chosen, *out[idx]):
                chosen.append(tuple(out[idx]))
        tps = sorted(chosen, key=lambda p: (near_gap(*p), ang(*p)))
        pins = sorted(range(n), key=lambda i: (near_gap(board.traces[i].start_x, board.traces[i].start_y),
                                               ang(board.traces[i].start_x, board.traces[i].start_y)))
    else:
        chosen = []
        for idx in np.argsort(-np.hypot(out[:, 0] - cx, out[:, 1] - cy)):
            if len(chosen) >= n:
                break
            if check_tp_spacing(chosen, *out[idx]):
                chosen.append(tuple(out[idx]))
        if spec.assignment == "gap_aware":
            tps = sorted(chosen, key=lambda p: (near_gap(*p), ang(*p)))
            pins = sorted(range(n), key=lambda i: (near_gap(board.traces[i].start_x, board.traces[i].start_y),
                                                   ang(board.traces[i].start_x, board.traces[i].start_y)))
        else:
            tps = sorted(chosen, key=lambda p: ang(*p))
            pins = sorted(range(n), key=lambda i: ang(board.traces[i].start_x, board.traces[i].start_y))

    placed = [None] * n
    for k, i in enumerate(pins):
        placed[i] = tps[k] if k < len(tps) else tps[-1]
    return board, placed


def challenge_board(board_size: float = 120.0, num_traces: int = 20, n_gaps: int = 3,
                    seed: int = 6, assignment: str = "angle", placement: str = "ring"):
    """Back-compat wrapper around make_challenge/ChallengeSpec."""
    return make_challenge(ChallengeSpec(
        board_size=board_size, num_traces=num_traces, n_gaps=n_gaps, seed=seed,
        assignment=assignment, placement=placement))


def _transform_xy(x: float, y: float, W: float, H: float,
                  quarter_turns: int, mirror: bool) -> Tuple[float, float]:
    """Optional x-mirror, then k*90deg CCW rotation, in a WxH board frame (origin 0,0)."""
    if mirror:
        x = W - x
    k = quarter_turns % 4
    if k == 1:
        return H - y, x
    if k == 2:
        return W - x, H - y
    if k == 3:
        return y, W - x
    return x, y


def rotate_board(board: BoardSpec, quarter_turns: int,
                 mirror: bool = False) -> BoardSpec:
    """Return a new BoardSpec rotated quarter_turns*90deg CCW (optionally x-mirrored first)."""
    W, H = board.width, board.height
    k = quarter_turns % 4
    nW, nH = (H, W) if k % 2 else (W, H)
    ox, oy = board.origin_x, board.origin_y
    nb = BoardSpec(origin_x=ox, origin_y=oy, width=nW, height=nH)

    def t(x, y):
        nx, ny = _transform_xy(x - ox, y - oy, W, H, k, mirror)
        return nx + ox, ny + oy

    for o in board.rect_obstacles:
        cx, cy = t(o.cx, o.cy)
        w, h = (o.height, o.width) if k % 2 else (o.width, o.height)
        nb.rect_obstacles.append(Obstacle(cx, cy, w, h, o.clearance, o.name))
    for o in board.circ_obstacles:
        cx, cy = t(o.cx, o.cy)
        nb.circ_obstacles.append(
            CircularObstacle(cx, cy, o.radius, o.clearance, o.name))
    ccx, ccy = t(board.connector_x + board.connector_w / 2.0,
                 board.connector_y + board.connector_h / 2.0)
    cw, ch = ((board.connector_h, board.connector_w) if k % 2
              else (board.connector_w, board.connector_h))
    nb.connector_x, nb.connector_y = ccx - cw / 2.0, ccy - ch / 2.0
    nb.connector_w, nb.connector_h = cw, ch
    for tr in board.traces:
        sx, sy = t(tr.start_x, tr.start_y)
        nb.traces.append(TraceSpec(start_x=sx, start_y=sy,
                                   breakout_length=tr.breakout_length,
                                   index=tr.index))
    return nb


def rotate_points(points, board: BoardSpec, quarter_turns: int,
                  mirror: bool = False):
    """Transform points with the same transform as rotate_board; None entries pass through."""
    ox, oy = board.origin_x, board.origin_y
    out = []
    for p in points:
        if p is None:
            out.append(None)
            continue
        nx, ny = _transform_xy(p[0] - ox, p[1] - oy, board.width, board.height,
                               quarter_turns % 4, mirror)
        out.append((nx + ox, ny + oy))
    return out


def generate_candidate_grid(board: BoardSpec, resolution: float = 6.5,
                            max_candidates: int = MAX_CANDIDATES
                            ) -> Tuple[np.ndarray, int]:
    """Valid TP candidates: (max_candidates, 2) array padded with (x_min, y_min) past real_count."""
    candidates = []

    x_lo = board.x_min + TP_TO_EDGE_MIN
    x_hi = board.x_max - TP_TO_EDGE_MIN
    y_lo = board.y_min + TP_TO_EDGE_MIN
    y_hi = board.y_max - TP_TO_EDGE_MIN

    xs = np.arange(x_lo, x_hi + resolution / 2, resolution)
    ys = np.arange(y_lo, y_hi + resolution / 2, resolution)

    for x in xs:
        for y in ys:
            if _is_valid_tp_position(board, x, y):
                candidates.append((x, y))

    # Uniform subsample preserves spatial coverage (no head-truncating one side).
    if len(candidates) > max_candidates:
        idx = np.linspace(0, len(candidates) - 1, max_candidates).astype(int)
        candidates = [candidates[i] for i in idx]

    real_count = len(candidates)

    while len(candidates) < max_candidates:
        candidates.append((board.x_min, board.y_min))

    return np.array(candidates, dtype=np.float64), real_count


def _is_valid_tp_position(board: BoardSpec, x: float, y: float) -> bool:
    """Check if (x, y) is a valid test point position."""
    if (x - board.x_min < TP_TO_EDGE_MIN or board.x_max - x < TP_TO_EDGE_MIN or
            y - board.y_min < TP_TO_EDGE_MIN or board.y_max - y < TP_TO_EDGE_MIN):
        return False

    conn_xmin = board.connector_x
    conn_xmax = board.connector_x + board.connector_w
    conn_ymin = board.connector_y
    conn_ymax = board.connector_y + board.connector_h
    dx = max(conn_xmin - x, 0, x - conn_xmax)
    dy = max(conn_ymin - y, 0, y - conn_ymax)
    if dx == 0 and dy == 0 and conn_xmin <= x <= conn_xmax and conn_ymin <= y <= conn_ymax:
        return False
    dist_to_conn = np.sqrt(dx**2 + dy**2)
    if dist_to_conn < TP_TO_CONNECTOR_MIN:
        return False

    for obs in board.rect_obstacles:
        xmin, ymin, xmax, ymax = obs.bounds
        buf = obs.clearance
        if xmin - buf < x < xmax + buf and ymin - buf < y < ymax + buf:
            return False

    for obs in board.circ_obstacles:
        dist = np.sqrt((x - obs.cx)**2 + (y - obs.cy)**2)
        if dist < obs.radius + obs.clearance:
            return False

    return True


def check_tp_spacing(placed_tps: List[Tuple[float, float]], x: float, y: float) -> bool:
    """Check if new TP at (x,y) satisfies spacing with all placed TPs."""
    for px, py in placed_tps:
        dist = np.sqrt((x - px)**2 + (y - py)**2)
        if dist < TP_TO_TP_MIN:
            return False
    return True